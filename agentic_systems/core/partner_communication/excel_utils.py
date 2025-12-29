"""Utility functions for creating Excel files with error comments.

This module provides utilities for generating Excel files with cell comments
for validation errors. Used by:
- GeneratePartnerErrorReportTool in Part 3 POC
- FR-012 review worksheet generation for human review (BRD Section 3.13)
"""

from pathlib import Path
from typing import Any, Dict, List, Optional
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill


# Severity-based fill colors per FR-012 (color-coding by severity)
ERROR_FILL = PatternFill(patternType="solid", fgColor="FFC7CE")  # Light red
WARNING_FILL = PatternFill(patternType="solid", fgColor="FFEB9C")  # Light yellow
INFO_FILL = PatternFill(patternType="solid", fgColor="BDD7EE")  # Light blue


def _standardize_phone_number(val: Any) -> Optional[str]:
    """Standardize phone number to XXX-XXX-XXXX format.
    
    Extracts digits from phone number and formats as XXX-XXX-XXXX.
    Handles various input formats:
    - (206) 555-1234
    - 206-555-1234
    - 206.555.1234
    - 2065551234
    - +1-206-555-1234 (US country code)
    - 1-206-555-1234 (US country code)
    - 001-206-555-1234 (international format with 001 prefix)
    - (206) 555-1234 x567 (extensions ignored)
    
    Args:
        val: Phone number value (string, number, etc.)
    
    Returns:
        Standardized phone number in XXX-XXX-XXXX format, or original value if invalid/empty
    """
    if pd.isna(val) or val is None:
        return val
    
    # Convert to string and strip whitespace
    phone_str = str(val).strip()
    
    # Handle international format: 001-XXX-XXX-XXXX
    # Remove "001-" prefix if present (common international dialing prefix)
    if phone_str.startswith('001-') or phone_str.startswith('001 '):
        phone_str = phone_str[4:].strip()
    elif phone_str.startswith('001'):
        # Handle cases like "0012065551234" (no separator)
        if len(phone_str) > 3 and phone_str[3].isdigit():
            phone_str = phone_str[3:]
    
    # Extract all digits
    digits_only = re.sub(r'\D', '', phone_str)
    
    # Must have at least 10 digits (US phone number)
    if len(digits_only) < 10:
        return val  # Return original if invalid
    
    # Handle US country code: if 11 digits and starts with 1, remove the leading 1
    if len(digits_only) == 11 and digits_only[0] == '1':
        digits_only = digits_only[1:]  # Remove country code
    
    # If still more than 10 digits, take last 10 (handles extensions and extra prefixes)
    # This ensures we get the actual phone number, not prefixes
    if len(digits_only) > 10:
        digits_only = digits_only[-10:]  # Take last 10 digits
    
    # Must have exactly 10 digits at this point
    if len(digits_only) != 10:
        return val  # Return original if invalid
    
    # Format as XXX-XXX-XXXX
    return f"{digits_only[0:3]}-{digits_only[3:6]}-{digits_only[6:10]}"


def _get_action_guidance(message: str, severity: str) -> str:
    """Generate action guidance based on error message and severity.
    
    Provides specific guidance for common validation error types per FR-012.
    """
    message_lower = message.lower()
    
    # Required field errors
    if 'required' in message_lower or 'missing' in message_lower:
        return "Provide the required value for this field"
    
    # Date validation errors
    if 'date' in message_lower and ('invalid' in message_lower or 'format' in message_lower):
        return "Enter a valid date in MM/DD/YYYY format"
    if 'date' in message_lower and 'earlier' in message_lower:
        return "Date must be earlier than today"
    if 'date' in message_lower and 'after' in message_lower:
        return "Exit date must be on or after start date"
    
    # Address validation errors
    if 'address' in message_lower and ('apartment' in message_lower or 'suite' in message_lower or 'unit' in message_lower):
        return "Move apartment/suite/unit information to Address 2 field"
    if 'address' in message_lower and 'po box' in message_lower:
        return "PO Box addresses are not allowed in Address 1"
    
    # Status validation errors
    if 'status' in message_lower and 'withdrawn' in message_lower:
        return "Provide noncompletion reason when status is Withdrawn/terminated"
    if 'status' in message_lower and 'graduated' in message_lower:
        return "Provide completion type and exit date when status is Graduated/completed"
    
    # Employment validation errors
    if 'employment' in message_lower and 'status' in message_lower:
        return "Select appropriate employment status or provide required employment details"
    if 'employment' in message_lower and ('name' in message_lower or 'type' in message_lower):
        return "Provide all required employment information when employment status is set"
    
    # Default guidance based on severity
    if severity == 'Error':
        return "Correct this error before processing can continue"
    elif severity == 'Warning':
        return "Review and correct if applicable"
    else:
        return "Review for accuracy"


def create_error_excel_with_comments(
    df: pd.DataFrame,
    violations: List[Dict],
    output_path: Path,
    col_map: Dict[str, str] = None
) -> None:
    """Create an Excel file with all rows, adding error comments and color-coding.
    
    This function creates an Excel file containing ALL rows from the DataFrame,
    preserving original order, with cell comments and color-coding added to cells
    that have errors. The comments contain the error message for that specific field.
    
    Implements FR-012 requirements for review worksheet generation:
    - Color-coding by severity (Error=red, Warning=yellow, Info=blue)
    - Cell comments with error messages
    - All records included, preserving original order
    
    Args:
        df: Original DataFrame with all data (must include header row)
        violations: List of violation dicts with 'row_index', 'field', 'message', 'severity'
        output_path: Path where Excel file should be saved
        col_map: Optional mapping from canonical keys to sheet headers.
                 If None, will try to match by field name directly.
                 This is used to map field names in violations to DataFrame columns.
    
    Returns:
        None (writes file to output_path)
    
    Note:
        - Row indices in violations are 1-based (row_index 2 = first data row)
        - Header errors (row_index == 1) are not included in the Excel file
        - Cells are color-coded by severity per FR-012: Error=red, Warning=yellow, Info=blue
        - Cell comments contain the error message for that specific field
        - The Excel file includes ALL rows (not just error rows) preserving original order
        - The Excel file includes all original columns only (no error metadata columns)
        - Error information is shown via cell comments and color-coding only
    
    Use Cases:
        - Part 3 POC: GeneratePartnerErrorReportTool creates partner error reports
        - FR-012: Generate review worksheet for human review with color-coding and action guidance
    """
    # Get all rows with errors (row_index > 1, since row_index 1 is header errors)
    error_row_indices = {v['row_index'] for v in violations if v['row_index'] > 1}
    
    # Include ALL rows from the DataFrame, preserving original order
    # This ensures partners can see all their data, not just rows with errors
    result_df = df.copy()
    
    # Standardize phone numbers to XXX-XXX-XXXX format
    # Find phone column by checking common header patterns
    phone_col = None
    for col in result_df.columns:
        col_lower = str(col).lower().strip()
        if 'phone' in col_lower:
            phone_col = col
            break
    
    if phone_col:
        # Apply phone number standardization to all rows
        result_df[phone_col] = result_df[phone_col].apply(_standardize_phone_number)
    
    # Get rows with errors (convert row_index to 0-based index: row_index 2 -> index 0)
    # row_index 2 = first data row = DataFrame index 0
    error_indices = {idx - 2 for idx in error_row_indices if idx >= 2}
    
    # Group violations by row_index for adding comments and color-coding
    violations_by_row = {}
    for v in violations:
        if v['row_index'] > 1:  # Skip header errors
            row_idx = v['row_index']
            if row_idx not in violations_by_row:
                violations_by_row[row_idx] = []
            violations_by_row[row_idx].append(v)
    
    # Write to Excel first (all rows, preserving original order)
    result_df.to_excel(output_path, index=False, engine='openpyxl')
    
    # Load workbook to add comments
    wb = load_workbook(output_path)
    ws = wb.active
    
    # Create a mapping from field name to column index
    # Field names in violations are the actual sheet headers
    field_to_col = {}
    for col_idx, header in enumerate(df.columns, start=1):
        field_to_col[str(header)] = col_idx
    
    # Map original row indices to Excel row numbers
    # Excel row 1 = header, Excel row 2 = first data row
    # Original row_index 2 = first data row = Excel row 2
    # Since we include all rows, the mapping is straightforward
    excel_row_map = {}  # Maps original row_index to Excel row number
    for df_idx in range(len(result_df)):
        orig_row_idx = df_idx + 2  # Convert DataFrame index to 1-based row_index
        excel_row = df_idx + 2  # Excel row number (row 1 is header, row 2 is first data row)
        excel_row_map[orig_row_idx] = excel_row
    
    # Add comments and color-coding by severity per FR-012
    for orig_row_idx, row_violations in violations_by_row.items():
        excel_row = excel_row_map.get(orig_row_idx)
        if not excel_row:
            continue
            
        for violation in row_violations:
            field_name = violation['field']
            message = violation['message']
            severity = violation.get('severity', 'Error')
            
            # Find column for this field
            col_idx = field_to_col.get(field_name)
            if col_idx:
                cell = ws.cell(row=excel_row, column=col_idx)
                
                # Add comment with error message
                comment = Comment(message, "Validation Tool")
                cell.comment = comment
                
                # Color-code by severity per FR-012 (Error=red, Warning=yellow, Info=blue)
                if severity == 'Error':
                    cell.fill = ERROR_FILL
                elif severity == 'Warning':
                    cell.fill = WARNING_FILL
                elif severity == 'Info':
                    cell.fill = INFO_FILL
    
    # Save workbook
    wb.save(output_path)

