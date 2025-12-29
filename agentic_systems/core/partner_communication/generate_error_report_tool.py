"""GeneratePartnerErrorReportTool per BRD FR-012.

Creates Excel file with full row data for rows containing validation errors.
Uses create_error_excel_with_comments for user-friendly Excel format with
color-coding and cell comments. This file contains PII and must be shared
via secure link only per BRD FR-013.

Includes WSAC aggregates worksheet per BRD FR-004 for partner review before
submission to WSAC.
"""

from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import load_workbook

from ..tools import ToolResult
from .excel_utils import create_error_excel_with_comments
from ...clients.cfa.rules import WRAPAROUND_SERVICE_NAMES


class GeneratePartnerErrorReportTool:
    """Generates error report Excel file with full row data for error rows per BRD FR-012.
    
    Takes validation violations and staged DataFrame, filters to error rows,
    and uses create_error_excel_with_comments to create user-friendly Excel file
    with color-coding, cell comments, and action guidance.
    
    Security Note: This Excel file contains full PII - must be shared via secure link only,
    never via email attachment per BRD FR-013 and BRD Section 2.3.
    """
    
    name = "GeneratePartnerErrorReportTool"
    
    def __call__(
        self,
        staged_dataframe: pd.DataFrame,
        violations: List[Dict[str, Any]],
        output_path: Path,
        aggregates: Optional[Dict[str, Any]] = None,
    ) -> ToolResult:
        """Generate error report Excel file with full row data for error rows.
        
        Uses create_error_excel_with_comments to create Excel file with:
        - Color-coding by severity (Error=red, Warning=yellow, Info=blue)
        - Cell comments with error messages
        - Action Required column with specific guidance
        - Notes column for additional context
        - WSAC Aggregates worksheet (if aggregates provided) per BRD FR-004
        
        Args:
            staged_dataframe: Original DataFrame from ingestion (with all columns)
            violations: List of violation dicts from ValidateStagedDataTool
                       Each dict has: row_index, field, severity, message
            output_path: Path where Excel file should be written (should be .xlsx)
            aggregates: Optional aggregates from CollectWSACAggregatesTool per BRD FR-004
        
        Returns:
            ToolResult with error_report_path in data
        """
        # Ensure output path is .xlsx
        if output_path.suffix != '.xlsx':
            output_path = output_path.with_suffix('.xlsx')
        
        # Extract unique row_index values from violations (filter out -1 for file-level errors)
        error_row_indices = {v['row_index'] for v in violations if v['row_index'] > 1}
        
        # Use create_error_excel_with_comments to generate Excel file with ALL rows
        # (not just error rows), preserving original order, with comments and color-coding
        # This provides a user-friendly format for partners to review and correct errors
        create_error_excel_with_comments(
            df=staged_dataframe,
            violations=violations,
            output_path=output_path
        )
        
        # Add WSAC Aggregates worksheet if aggregates provided per BRD FR-004
        if aggregates:
            self._add_aggregates_worksheet(output_path, aggregates)
        
        # Count error rows and total rows
        error_row_count = len(error_row_indices)
        total_row_count = len(staged_dataframe)
        
        return ToolResult(
            ok=True,
            summary=f"Generated error report Excel file with {total_row_count} total rows ({error_row_count} with errors)",
            data={
                "error_report_path": str(output_path),
                "error_row_count": error_row_count,
                "total_row_count": total_row_count
            },
            warnings=[],
            blockers=[]
        )
    
    def _add_aggregates_worksheet(self, workbook_path: Path, aggregates: Dict[str, Any]) -> None:
        """Add Quarterly Updates worksheet to error report per BRD FR-004.
        
        Creates a separate worksheet with three sections:
        1. Number of participants using each wraparound service
        2. Amount of GJC funds spent on each service type
        3. Amount of non-GJC funds spent on each service type
        
        Args:
            workbook_path: Path to Excel workbook
            aggregates: Aggregates dictionary from CollectWSACAggregatesTool
        """
        wb = load_workbook(workbook_path)
        
        # Create new worksheet for aggregates
        ws_aggregates = wb.create_sheet("Quarterly Updates")
        
        # Get wraparound services data
        wraparound = aggregates.get('wraparound_services', {})
        usage_counts = wraparound.get('usage_counts', {})
        gjc_funds = wraparound.get('gjc_funds_spent', {})
        non_gjc_funds = wraparound.get('non_gjc_funds_spent', {})
        
        # Define service types in order (matching rules.py)
        service_types = [
            'transportation',
            'childcare',
            'career_services_and_learning_materials',
            'mental_health_services',
            'life_skills',
            'navigation',
            'other'
        ]
        
        # Section 1: How many GJC Participants used each type of wraparound service
        ws_aggregates.append(["How many GJC Participants used each type of wraparound service in the past quarter?"])
        ws_aggregates.append(["Service Type", "Number of Participants"])
        
        for service_type in service_types:
            service_name = WRAPAROUND_SERVICE_NAMES.get(service_type, service_type.replace('_', ' ').title())
            count = usage_counts.get(service_type, 0)
            ws_aggregates.append([service_name, count])
        
        # Add row for "other" specification
        ws_aggregates.append(["If there are \"other\" wraparound services, please specify", ""])
        
        # Empty row between sections
        ws_aggregates.append([])
        
        # Section 2: What was the amount of GJC funds spent on each type of service?
        ws_aggregates.append(["What was the amount of GJC funds spent on each type of service?"])
        ws_aggregates.append(["Service Type", "Amount of GJC funds"])
        
        for service_type in service_types:
            service_name = WRAPAROUND_SERVICE_NAMES.get(service_type, service_type.replace('_', ' ').title())
            gjc = gjc_funds.get(service_type) if gjc_funds.get(service_type) is not None else None
            # Format as currency if value exists, otherwise leave blank
            gjc_value = f"${gjc:,.2f}" if gjc is not None and gjc != 0 else ""
            ws_aggregates.append([service_name, gjc_value])
        
        # Add row for "other" specification
        ws_aggregates.append(["If there are \"other\" wraparound services, please specify", ""])
        
        # Empty row between sections
        ws_aggregates.append([])
        
        # Section 3: What was the amount of non-GJC funds spent on each type of service?
        ws_aggregates.append(["What was the amount of non-GJC funds spent on each type of service?"])
        ws_aggregates.append(["Service Type", "Amount of non-GJC funds"])
        
        for service_type in service_types:
            service_name = WRAPAROUND_SERVICE_NAMES.get(service_type, service_type.replace('_', ' ').title())
            non_gjc = non_gjc_funds.get(service_type) if non_gjc_funds.get(service_type) is not None else None
            # Format as currency if value exists, otherwise leave blank
            non_gjc_value = f"${non_gjc:,.2f}" if non_gjc is not None and non_gjc != 0 else ""
            ws_aggregates.append([service_name, non_gjc_value])
        
        # Add row for "other" specification
        ws_aggregates.append(["If there are \"other\" wraparound services, please specify", ""])
        
        wb.save(workbook_path)

